"""use openpyxl to open an excel sheet

use python classes to sort attributes of the excel sheet
"""

import csv
import openpyxl


def csv_to_excel(filename):
    """converts csv file to xlsx using openpyxl"""
    csv_data = []
    with open(filename) as file_obj:
        reader = csv.reader(file_obj)
        for row in reader:
            csv_data.append(row)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = filename[:-4]
    for row in csv_data:
        sheet.append(row)
    return workbook


class VirpytWorkbook():
    """workbook class with pointer to the workbook

    constructor/attribute __init__() that takes a workbookname
    and then uses openpyxl to open that workbook and store the
    workbook pointer in a local member.
    """

    def __init__(self, filename):
        """opens workbook using openpyxl and stores pointer locally"""
        if filename[-3:] == "csv":
            self._workbook = csv_to_excel(filename)

        else:
            self._workbook = openpyxl.load_workbook(filename)

        self._sheet_map = {}

        #wraps openpyxl sheet objects using Sheet class
        for sheetname in self.sheetnames:
            self._sheet_map[sheetname] = VirpytSheet(self._workbook[sheetname],
                                               sheetname)


    @property
    def sheetnames(self):
        """returns a list of sheetnames"""

        return self._workbook.sheetnames


    @property
    def sheets(self):
        """returns a list of VirPyTSheet"""
        return [self._sheet_map[sheetname] for sheetname in self._sheet_map]



class VirpytSheet():
    """sheet class with pointers to sheets in the workbook"""

    def __init__(self, sheet, name):
        self._sheet = sheet     #openpyxl sheet obj
        self._name = name
        self._tables = {}       #{startcell : values}


    @property
    def name(self):
        """returns sheetname"""
        return self._name


    #table objects
    @property
    def tables(self):
        """finds and stores tables"""
        self.find_tables()
        return [v for _,v in self._tables.items()]


    def find_tables(self):
        """finds tables

        using openpyxl to find the first cell with data,
        the next empty column and row are found and tracked to show
        dimensions of the table and location on the worksheet,
        then method startcell() is called to find the next table, if any.
        """
        # store worksheet starting cell indices
        tablestartrow = self._sheet.min_row
        tablestartcol = self._sheet.min_column

        #find first cell with values using min_row, min_column
        startcell = self._sheet.cell(row=self._sheet.min_row,
                                     column=self._sheet.min_column).coordinate

        # loop to find table dimensions
        while tablestartrow < self._sheet.max_row:

            # 0-based numrow & numcol
            # thesse will store dimensions of table
            numrow = 0
            numcol = 0

            # scan until empty column is found
            for col in self._sheet.iter_cols(min_col=tablestartcol,
                                             min_row=tablestartrow):
                if col[0].value:    #header row should extend over whole table
                    numcol += 1     #count columns in header row
                else:
                    break           #break when no value in header row

            # scan until empty row is found
            for row in self._sheet.iter_rows(min_row=tablestartrow,
                                             min_col=tablestartcol,
                                             max_col=tablestartcol+numcol-1):
                #sometimes theres a gap where one column doesnt have data
                emptyrow = True
                for cell in row:
                    if cell.value:  #rows with values will increment numrow
                        emptyrow = False
                if emptyrow:        #rows without values will break
                    break
                else:
                    numrow += 1

            # store table starting cell coords as tuple
            coords = (tablestartcol, tablestartrow)

            # store table ending cell indices
            tableendcol = tablestartcol+numcol-1
            tableendrow = tablestartrow+numrow-1

            # store table information in _tables dict
            self._tables[startcell] = VirpytTable(self._sheet, coords,
                                                  numcol, numrow)

            # use ending cell indices to find next table starting cell
            startcell, tablestartcol, tablestartrow = self.startcell(
                                                      tableendcol,
                                                      tableendrow,
                                                      numcol,
                                                      numrow)


    def startcell(self, tableendcol, tableendrow, numcol, numrow):
        """finds next table's starting location on worksheet"""

        # search vertically for next table
        #assume tables are aligned, check first col for empty cells
        for row in self._sheet.iter_rows(min_row=tableendrow):
            # count empty rows by checking below
            # first column of previous table
            if not row[self._sheet.min_row].value:
                tableendrow +=1

        """
        # start seraching horizontally if maxrows reached
        if tableendrow == self._sheet.max_row:
            for col in self._sheet.iter_cols(min_col = tableendcol):
                if not col[0].value:
                    tableendcol += 1

            return self._sheet.cell(row=(tableendrow-numrow+1),
                                    column=tableendcol).coordinate
        else:
            return self._sheet.cell(row=tableendrow,
                                    column=(tableendcol-numcol+1)).coordinate
        """

        startcell = self._sheet.cell(column=(tableendcol-numcol+1),
                                     row=tableendrow)
        return startcell, startcell.column, startcell.row
        #scan til cell with data, use as startcell for next table



class VirpytTable():
    """table objects hold ranges of cells on worksheet"""

    def __init__(self, sheet, coords, numcol, numrow):

        #defines table object with starting cell and dimensions
        self._sheet = sheet

        self._coords = coords
        self._numcol = numcol
        self._numrow = numrow
        self._rows = []
        self._colmap = {}


        self.header = VirpytRow(self, 0).rowvals

        #
        for item in self.header:
            self._colmap[item] = self.header.index(item)


    @property
    def rows(self):
        """wraps rows in Row objects"""
        return [VirpytRow(self, index) for index in range(self._numrow)]


    @property
    def columns(self):
        """wraps columns in Column objects"""
        return [VirpytColumn(self, index) for index in range(self._numcol)]


    @property
    def coords(self):
        """returns table's starting cell coordinates"""
        return self._coords



class VirpytRow():
    """Row objects hold horizontal values of cells in table"""

    def __init__(self, table, index):
        self._table = table
        self._index = index


    @property
    def rowvals(self):
        """returns cell values by row"""

        sheet = self._table._sheet

        y = self._index + self._table.coords[1]

        # cell coordinates are stored in openpyxl as (row, column)
        return [sheet.cell(y,x).value for x in range(self._table._coords[0],
                       self._table._numcol + self._table._coords[0])]

    # use cell name to return index of table header
    def __getitem__(self, key):

        # get column index per key
        col_index = self._table._colmap[key]
        sheet = self._table._sheet
        coords = self._table._coords

        return sheet.cell(coords[0]+self._index, coords[1]+col_index)




class VirpytColumn():
    """Column objects hold vertical values of cells in table"""

    def __init__(self, table, index):
        self._table = table
        self._index = index

    @property
    def colvals(self):
        """returns cell values by column"""

        sheet = self._table._sheet

        x = self._index + self._table.coords[0]

        retval = []
        for y in range(self._table._coords[1],
                       self._table._numrow + self._table._coords[1]):
            cellval = sheet.cell(y,x).value
            retval.append(cellval)

        return retval
