# use openpyxl to open an excel sheet
# use python classes to sort attributes of the excel sheet


import openpyxl


# workbook class with pointer to the workbook
class VirpytWorkbook():
    
    # constructor/attribute __init__() that takes a workbookname
    # and then uses openpyxl to open that workbook and store the
    # workbook pointer in a local member.  
    def __init__(self, filename):
        self._workbook = openpyxl.load_workbook(filename)
    
    
    # .sheetnames returns a list of sheetnames
    @property
    def sheetnames(self):
        return self._workbook.sheetnames
    
    
    # returns a list of VirPyTSheet
    @property
    def sheets(self):
                #wraps openpyxl sheet objects using Sheet class
        return [Sheet(self._workbook[sheetname], sheetname)
                for sheetname in self.sheetnames]
    
    
    def save(self):
        self._workbook.save()



# sheet class with pointers to sheets in the workbook
class Sheet():
    
    def __init__(self, sheet, name):
        self._sheet = sheet     #openpyxl sheet obj
        self._name = name
        self._tables = {}       #{startcell : values}
    
    
    @property
    def name(self):
        return self._name
    
    #table objects
    @property
    def tables(self):
        self.find_tables()
        return [v for _,v in self._tables.items()]
    
    
    
    def find_tables(self):
        
        # store worksheet starting cell indices
        tablestartrow = self._sheet.min_row
        tablestartcol = self._sheet.min_column
        
        #find first cell with values using min_row, min_column
        startcell = self._sheet.cell(row=self._sheet.min_row,
                                     column=self._sheet.min_column).coordinate
        
        while tablestartrow < self._sheet.max_row:
            # 0-based numrow & numcol
            # thesse will show dimensions of table
            numrow = 0
            numcol = 0
            
            # scan until empty column is found
            for col in self._sheet.iter_cols(min_col=tablestartcol,
                                             min_row=tablestartrow):
                if col[0].value:    #header row should extend over whole table
                    numcol += 1     #count columns in header row
                else:
                    break           #break when no value in header row
            
            # scan until empty row is found
            for row in self._sheet.iter_rows(min_row=tablestartrow,
                                             min_col=tablestartcol,
                                             max_col=tablestartcol+numcol-1):
                #sometimes theres a gap where one column doesnt have data
                emptyrow = True
                for cell in row:
                    if cell.value:  #rows with values will increment numrow
                        emptyrow = False
                if emptyrow:        #rows without values will break
                    break
                else:
                    numrow += 1
            
            # store table starting cell coords as tuple
            coords = (tablestartcol, tablestartrow)
            
            # store table ending cell indices
            tableendcol = (tablestartcol + numcol -1)
            tableendrow = (tablestartrow + numrow -1)
            
            # store table information in _tables dict
            self._tables[startcell] = Table(self._sheet, coords,
                                            numcol, numrow)
            
            # use ending cell indices to find next table starting cell
            startcell, tablestartcol, tablestartrow = self.startcell(
                                                      tableendcol,
                                                      tableendrow,
                                                      numcol,
                                                      numrow)


    def startcell(self, tableendcol, tableendrow, numcol, numrow):
        
        # search vertically for next table
        #assume tables are aligned, check first col for empty cells
        for row in self._sheet.iter_rows(min_row=tableendrow):
            # count empty rows by checking below
            # first column of previous table
            if not row[self._sheet.min_row].value:
                tableendrow +=1
        
        '''
        # start seraching horizontally if maxrows reached
        if tableendrow == self._sheet.max_row:
            for col in self._sheet.iter_cols(min_col = tableendcol):
                if not col[0].value:
                    tableendcol += 1

            return self._sheet.cell(row=(tableendrow-numrow+1),
                                    column=tableendcol).coordinate
        else:
            return self._sheet.cell(row=tableendrow,
                                    column=(tableendcol-numcol+1)).coordinate
        '''
        
        startcell = self._sheet.cell(column=(tableendcol-numcol+1),
                                     row=tableendrow)
        return startcell, startcell.column, startcell.row
        #scan til cell with data, use as startcell for next table



class Table():
    def __init__(self, sheet, coords, numcol, numrow):
        
        #defines table object with starting cell and dimensions
        self._sheet = sheet
        
        self._coords = coords
        self._numcol = numcol
        self._numrow = numrow
        self._rows = []

    
    @property
    def header(self):
        return Row(self, 0)

    @property
    def rows(self):
        return [Row(self, index) for index in range(self._numrow)]
        

    @property
    def columns(self):
        return [Column(self, index) for index in range(self._numcol)]


    @property
    def coords(self):
        return self._coords
    


class Row():
    
    def __init__(self, table, index):
        self._table = table
        self._index = index

    @property
    def rowvals(self):
        sheet = self._table._sheet
        
        y = self._index + self._table.coords[1]

        '''
        retval = []
        for x in range(self._table._coords[0],
                       self._table.numcol + self._table._coords[0]):
            cellval = sheet.cell(x,y).value
            retval.append(cellval)
        return retval
        '''
        # cell coordinates are stored in openpyxl as (row, column)
        return [sheet.cell(y,x).value for x in range(self._table._coords[0],
                       self._table._numcol + self._table._coords[0])]
    
    
    def __getitem__(self, key):
        pass




class Column():
    
    def __init__(self, table, index):
        self._table = table
        self._index = index

    @property
    def colvals(self):
        sheet = self._table._sheet
        
        x = self._index + self._table.coords[0]
        
        retval = []
        for y in range(self._table._coords[1],
                       self._table._numrow + self._table._coords[1]):
            cellval = sheet.cell(y,x).value
            retval.append(cellval)
        
        return retval
    
    
    def __getitem__(self, key):
        pass
    
            
#class Cell():

              

if __name__ == '__main__':
    #workbookname = input(print("Input workbookname: "))
    
    wb = VirpytWorkbook('sample.xlsx')
    print(wb)        #address of openpyxl workbook object

    print(wb.sheets)        #list of sheet object addresses
    
    print(wb.sheetnames)    #list of names of worksheets

    for sheet in wb.sheets:     #prints each sheet title
        print("Found sheet named %s" % sheet.name)
        for table in sheet.tables:
            print("Found table: ", table.coords,
                  table._numcol, table._numrow)
            for row in table.rows:
                print(row.rowvals)
            for column in table.columns:
                print(column.colvals)
