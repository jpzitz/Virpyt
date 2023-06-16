import openpyxl
#from openpyxl import Workbook







wb = openpyxl.load_workbook(input("Input filename: "))

# workbook class with pointer to the workbook
class VirPyTWorkbook:
    def __init__(self, file):
        self._file = file
        
    @property
    def file(self):
        """The file property."""
        print("Get file")
        return self._file

    @file.setter
    def file(self, filename):
        print("Set filename")
        self._file = filename

    @file.deleter
    def file(self):
        print("Clear filename")
        del self._file
    
    sheets = []



# sheet class with pointers to sheets in the workbook        
#class VirPyTSheet():
    

    
#class VirPyTTable():
#class VirPyTRow():
#class VirPyTCell():
  
#worksheet = workbook.active
#print(wb.sheetnames)

    

workbook = VirPyTWorkbook(_wb)
print(workbook._file)

worksheet = wb.active
if worksheet:
    print("ok!")

for sheet in workbook.sheets:
    print("Found sheet named %s" %sheet.title)
    
    for table in sheet.tables:
            print("Found table named %t" %table.title)

     #   for column_header in table.header:  # iterate through cells that make
                                            # up the header
      #  print(“col header is %s” % column_header.value)
       # for row in table:
        #    cell = row['<col name>']        # we want to be able to access a
                                            # cell according to the header name.
         #   print('row info is %s' % cell.value)
