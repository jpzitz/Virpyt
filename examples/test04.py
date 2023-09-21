import openpyxl

class Book():
    def __init__(self, file):
        print(file[-3:])
        self.file = openpyxl.load_workbook(file)

    @property
    def worksheets(self):
        return self.file.worksheets

    

wb = Book('sample.xlsx')
print(wb.file)
print(wb.worksheets)    #list of sheetnames
ws = wb.worksheets[0]
print(ws)


# csv_to_excel.py
import csv
import openpyxl
def csv_to_excel(csv_file, excel_file):
    csv_data = []
    with open(csv_file) as file_obj:
        reader = csv.reader(file_obj)
        for row in reader:
            csv_data.append(row)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row in csv_data:
        sheet.append(row)
    workbook.save(excel_file)
